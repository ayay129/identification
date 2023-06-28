#! /usr/bin/env python3
# coding: utf-8
# -*- v_xbinyan Create By 20180418 -*-

import xlrd
import openpyxl
# import chardet
import hashlib
import os
import threading
from multiprocessing.pool import ThreadPool
import io
import gzip
import zipfile
import time
import datetime
import re
import sys
# from selenium import webdriver

types = sys.getdefaultencoding()
locker = threading.Lock()
TIMEFORMATE = '%Y-%m-%d %H:%M:%S.%f'


# 脚本工具
class Tools(object):
    def currenttime(self, format=TIMEFORMATE):
        """获取当前时间

        :return: 返回当前时间，默认格式%Y-%m-%d %H:%M:%S.%f
        """
        return time.strftime(format, time.localtime())

    def getfilecharset(self, file):
        """获取文件编码

        :param file: 文件路径
        :return: 字符编码
        """
        r = chardet.detect(open(file).read())
        charenc = r['encoding']
        return charenc

    def readfile(self, file, charset='utf8'):
        """读取文件内容

        :param file: 文件地址
        :param charset: 文件编码
        :return: utf8编码的内容生成器
        """
        # charset = getfilecharset(file)
        with io.open(file, 'r', encoding=charset) as f:
            for line in f:
                yield line.replace('\n', '').strip()

    def write2file(self, file, lines):
        """将内容写入文件

        :param file: 输出文件地址
        :param lines: 写入内容
        :return:
        """
        with io.open(file, "w") as f:
            f.writelines([u''.join(line + '\n') for line in lines])

    def simple2tradition(self, line):
        """简体转繁体

        :param line: 输入内容，需要utf8编码
        :return: 转换后的繁体内容
        """
        # 将简体转换成繁体
        line = Converter('zh-hant').convert(line.decode('utf-8'))
        line = line.encode('utf-8')
        return line

    def tradition2simple(self, line):
        """繁体转简体

        :param line: 输入内容
        :return: 转换后的简体内容
        """
        # 将繁体转换成简体
        line = Converter('zh-hans').convert(line.decode('utf-8'))
        line = line.encode('utf-8')
        return line

    def get_md5_value(self, src):
        """计算md5

        :param src: 输入字符串
        :return: 字符串md5
        """
        myMd5 = hashlib.md5()
        myMd5.update(src)
        myMd5_Digest = myMd5.hexdigest()
        return myMd5_Digest

    def get_sha1_value(self, src):
        """计算sha1

        :param src: 输入字符串
        :return: 字符串sha1
        """
        mySha1 = hashlib.sha1()
        mySha1.update(src)
        mySha1_Digest = mySha1.hexdigest()
        return mySha1_Digest

    def getdirsindir(self, rootdir):
        """获取一级子目录

        获取一级子目录
        :param rootdir: 根目录
        :return: 子目录名称
        """
        for root, dirs, files in os.walk(rootdir):
            for dir in dirs:
                yield "%s/%s" % (root, dir)

    def getfilesindir(self, dir):
        """获取dir下所有文件

        :param dir: 输入文件夹地址
        :return: dir下所有文件路径
        """
        for root, dirs, files in os.walk(dir):
            for name in files:
                yield '%s/%s' % (root, name)
                # yield name

    def multiprocesstask(self, poolsize, func, iterator):
        """多线程执行任务

        :param poolsize: 线程池大小
        :param func: 执行函数
        :param iterator: 迭代器
        :return:
        """
        pool = ThreadPool(poolsize)
        pool.map(func=func, iterable=iterator)
        pool.close()
        pool.join()
        # return result

    def readdatafromtable(self, entity):
        """读取数据库数据

        :param entity: 表对应的entity
        :return:
        """
        return entity.objects.all()

    def savedata2table(self, entity, results):
        """保存数据到数据库

        :param entity: 表对应的entity
        :param results: 保存的结果
        :return:
        """
        while not locker.acquire(60):
            print('get locker failed')
        entity.objects.bulk_create(results)
        locker.release()

    def un_gz(self, file_name):
        """解压gz格式文件

        :param file_name: 文件路径
        :return:
        """
        f_name = file_name.replace(".gz", "")
        # 获取文件的名称，去掉
        g_file = gzip.GzipFile(file_name)
        # 创建gzip对象
        open(f_name, "w+").write(g_file.read())
        # gzip对象用read()打开后，写入open()建立的文件中。
        g_file.close()
        # 关闭gzip对象

    def un_zip(self, file_name, extractdir=''):
        """解压zip文件

        :param file_name:文件路径
        :return:
        """
        zip_file = zipfile.ZipFile(file_name)
        if '' == extractdir:
            if os.path.isdir(file_name.replace('.zip', '')):
                extractdir = file_name.replace('.zip', '')
            else:
                extractdir = os.mkdir(file_name.replace('.zip', ''))
        else:
            if os.path.isdir(extractdir):
                pass
            else:
                os.makedirs(extractdir)
        for names in zip_file.namelist():
            try:
                zip_file.extract(names, extractdir)
            except:
                pass

    def un_rar(self, file_name, extractdir=''):
        """unzip rar file"""
        rar = rarfile.RarFile(file_name)
        if '' == extractdir:
            if os.path.isdir(file_name.replace('.zip', '')):
                extractdir = file_name.replace('.zip', '')
            else:
                extractdir = os.mkdir(file_name.replace('.zip', ''))
        else:
            if os.path.isdir(extractdir):
                pass
            else:
                os.makedirs(extractdir)
        os.chdir(extractdir)
        rar.extractall()
        rar.close()

    def identifylanguage(self, str):
        """识别文本字符

        :param str: 输入字符串
        :return: 识别结果 1.jpg：中文 2：日语 3：韩语
        """
        # 日文
        re_jp = re.compile(u"[\u3040-\u309f\u30a0-\u30ff]+")
        # 韩文
        re_kr = re.compile(u"[\uac00-\ud7ff]+")
        match = re_jp.search(str, 0)
        if match:
            # 日语
            return 2
        match = re_kr.search(str, 0)
        if match:
            # 韩语
            return 3
        re_cn = re.compile(u"[\u4e00-\u9fa5]+")
        match = re_cn.search(str, 0)
        if match:
            # 中文
            return 1
        # 中文
        return -1

    def timestamp_to_strtime(self, timestamp, formate=TIMEFORMATE):
        """将 13 位整数的毫秒时间戳转化成本地普通时间 (字符串格式)

        :param timestamp: 13 位整数的毫秒时间戳 (1456402864242)
        :param formate: 字符串格式，默认'%Y-%m-%d %H:%M:%S.%f'
        :return: 返回字符串格式 {str}'2016-02-25 20:21:04.242000'
        """

        local_str_time = datetime.fromtimestamp(timestamp / 1000.0).strftime(formate)
        return local_str_time

    def timestamp_to_datetime(self, timestamp):
        """将 13 位整数的毫秒时间戳转化成本地普通时间 (datetime 格式)

        :param timestamp: 13 位整数的毫秒时间戳 (1456402864242)
        :return: 返回 datetime 格式 {datetime}2016-02-25 20:21:04.242000
        """
        local_dt_time = datetime.fromtimestamp(timestamp / 1000.0)
        return local_dt_time

    def datetime_to_strtime(self, datetime_obj, formate=TIMEFORMATE):
        """将 datetime 格式的时间 (含毫秒) 转为字符串格式

        :param datetime_obj: {datetime}2016-02-25 20:21:04.242000
        :param formate: 字符串格式，默认'%Y-%m-%d %H:%M:%S.%f'
        :return: {str}'2016-02-25 20:21:04.242'
        """
        local_str_time = datetime_obj.strftime(formate)
        return local_str_time

    def datetime_to_timestamp(self, datetime_obj):
        """将本地(local) datetime 格式的时间 (含毫秒) 转为毫秒时间戳

        :param datetime_obj: {datetime}2016-02-25 20:21:04.242000
        :return: 13 位的毫秒时间戳  1456402864242
        """
        local_timestamp = long(time.mktime(datetime_obj.timetuple()) * 1000.0 + datetime_obj.microsecond / 1000.0)
        return local_timestamp

    def strtime_to_datetime(self, timestr, formate=TIMEFORMATE):
        """将字符串格式的时间 (含毫秒) 转为 datetiem 格式

        :param timestr: {str}'2016-02-25 20:21:04.242'
        :param formate: 字符串格式，默认'%Y-%m-%d %H:%M:%S.%f'
        :return: {datetime}2016-02-25 20:21:04.242000
        """
        local_datetime = datetime.strptime(timestr, formate)
        return local_datetime

    def strtime_to_timestamp(self, local_timestr, formate=TIMEFORMATE):
        """将本地时间 (字符串格式，含毫秒) 转为 13 位整数的毫秒时间戳

        :param local_timestr: {str}'2016-02-25 20:21:04.242'
        :param formate: 字符串格式，默认'%Y-%m-%d %H:%M:%S.%f'
        :return: 1456402864242
        """
        local_datetime = strtime_to_datetime(local_timestr, formate)
        timestamp = datetime_to_timestamp(local_datetime)
        return timestamp

    def is_number(self, uchar):
        """判断一个unicode是否是数字

        :param uchar: 输入字符
        :return:
        """
        if u'\u0030' <= uchar <= u'\u0039':
            return True
        else:
            return False

    def is_alphabet(self, uchar):
        """判断一个unicode是否是英文字母

        :param uchar: 输入字符
        :return:
        """
        if (u'\u0041' <= uchar <= u'\u005a') or (u'\u0061' <= uchar <= u'\u007a'):
            return True
        else:
            return False

    def containdigitoralpha(self, str):
        """判断字符串包含数字或英文

        :param str: 输入字符串
        :return: True：包含 False：不包含
        """
        return any(is_number(char) or is_alphabet(char) for char in str)


# 前缀树搜索
class Trie(object):
    """前缀树

    构建前缀树，用于字符串搜索
    """
    root = {}
    END = '/'

    def __init__(self):
        """初始化函数
        """
        super(Trie, self).__init__()

    def add(self, word):
        """添加word
        :param word: 输入word
        :return:
        """
        # 从根节点遍历单词,char by char,如果不存在则新增,最后加上一个单词结束标志
        node = self.root
        for c in word:
            node = node.setdefault(c, {})
        node[self.END] = None

    def find(self, word):
        """搜索word
        :param word: 输入word
        :return: 搜索结果 True：存在 False：不存在
        """
        node = self.root
        for c in word:
            if c not in node:
                return False
            node = node[c]
        return self.END in node


# 读取工具类EXCEL
class analysis(object):
    """
    读取工具类
    """

    def __init__(self, path="file.xlsx"):
        self.filepath = path

    def open_excel(self, file='file.xlsx'):
        try:
            data = xlrd.open_workbook(file)
            return data
        except Exception:
            print("打开文件异常")

    def get_headers(self,file):
        xl = self.open_excel(file)
        return xl.sheet_names()
    def excel_table_byindex(self, file, colnameindex=0, by_index=0, start_row=1):
        data = self.open_excel(file)
        table = data.sheets()[by_index]
        nrows = table.nrows  # 行数
        ncols = table.ncols  # 列数
        colnames = table.row_values(colnameindex)  # 某一行数据
        list = []
        for rownum in range(start_row, nrows):
            row = table.row_values(rownum)
            if row:
                app = {}
                for i in range(len(colnames)):
                    app[colnames[i]] = row[i]
                list.append(app)
        return list

    def excel_table_byname(self, file, colnameindex=0, by_name='Sheet1', start_row=1):
        data = self.open_excel(file)
        table = data.sheet_by_name(by_name)
        nrows = table.nrows
        colnames = table.row_values(colnameindex)
        list = []
        for rownum in range(start_row, nrows):
            row = table.row_values(rownum)
            if row:
                app = {}
                for i in range(len(colnames)):
                    app[colnames[i]] = row[i]
                list.append(app)
        return list

    def excel_table_byrow(self, file, by_index=0, start_row=0):
        data = self.open_excel(file)
        table = data.sheets()[by_index]
        list = []
        nrows = table.nrows
        for rownum in range(start_row, nrows):
            row = table.row_values(rownum)
            if row:
                list.append(row)
        return list

    def openpy_excel_bycol(self, file, by_row=1, colnameindex=1):
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        max_rows = ws.max_row
        max_cols = ws.max_column
        list = []
        for col_index in range(colnameindex, max_cols + 1):
            obj_list = []
            for row_index in range(by_row, max_rows + 1):
                lattice = ws.cell(row=row_index, column=col_index)
                obj_list.append((lattice.value, lattice.fill.start_color.rgb))
            list.append(obj_list)
        return list

    def openpy_excel_byrow(self, file, by_row=1, colnameindex=1):
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        max_rows = ws.max_row
        max_cols = ws.max_column
        list = []
        for row_index in range(by_row, max_rows + 1):
            obj_list = []
            for col_index in range(colnameindex, max_cols + 1):
                lattice = ws.cell(row=row_index, column=col_index)
                color = lattice.fill.start_color.rgb
                obj_list.append((lattice.value, color))
            list.append(obj_list)
        return list

    def week_get(self, vdate):
        dayscount = datetime.timedelta(days=vdate.isoweekday())
        dayfrom = vdate - dayscount + datetime.timedelta(days=1)
        dayto = vdate - dayscount + datetime.timedelta(days=7)
        print(' ~~ '.join([str(dayfrom), str(dayto)]))
        week7 = []
        i = 0
        while (i <= 6):
            week7.append('周' + str(i + 1) + ': ' + str(dayfrom + datetime.timedelta(days=i)))
            i += 1
        return week7

    def weekInYear(self, vdate_str):
        date = vdate_str
        yearWeek = datetime.date(int(date[0:4]), int(date[5:7]), int(date[8:10])).isocalendar()[0:2]
        return str(yearWeek[0]) + '#' + str(yearWeek[1])
